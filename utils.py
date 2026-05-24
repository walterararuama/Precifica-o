import pandas as pd
import re
import os
import time
import shutil
import tempfile
from datetime import datetime
import tkinter as tk
import openpyxl
import logging

log = logging.getLogger(__name__)

def converter_moeda(valor_str):
    if pd.isna(valor_str) or valor_str == "": return 0.0
    v = str(valor_str).strip().replace('R$', '').replace(' ', '')
    if ',' in v and '.' in v: v = v.replace('.', '').replace(',', '.')
    elif ',' in v: v = v.replace(',', '.')
    try: return float(v)
    except: return 0.0

def formatar_moeda(valor): 
    return f"R$ {valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def formatar_percentual(valor): 
    return f"{valor:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

def auto_selecionar(event): 
    event.widget.after(10, lambda: event.widget.select_range(0, tk.END))

def arredondar_preco(valor_sugerido, custo_base=0.0, markup_alvo=0.0):
    if valor_sugerido <= 0: return 0.0
    base_centena = (int(valor_sugerido) // 100) * 100
    opcoes = []
    finais_permitidos = [19, 29, 39, 49, 59, 69, 79, 89, 99] if valor_sugerido < 300 else [49, 99]
    for i in finais_permitidos: opcoes.extend([base_centena + i, base_centena - 100 + i, base_centena + 100 + i])
    opcoes = sorted(list(set(opcoes)))
    melhor_opcao = min(opcoes, key=lambda x: abs(x - valor_sugerido))
    
    if melhor_opcao < valor_sugerido:
        pode_arredondar_baixo = False
        if (valor_sugerido - melhor_opcao) <= 11.5:
            pode_arredondar_baixo = True
            if custo_base > 0 and markup_alvo >= 0.90:
                if ((melhor_opcao / custo_base) - 1) < (markup_alvo - 0.05): pode_arredondar_baixo = False
        if not pode_arredondar_baixo:
            opcoes_acima = [x for x in opcoes if x >= valor_sugerido]
            melhor_opcao = min(opcoes_acima) if opcoes_acima else valor_sugerido
    return float(max(melhor_opcao, 19.0))

def check_nota_duplicada(nota_digitada, pasta_fretes):
    if not nota_digitada: return False
    nota_limpa = re.sub(r'\.0$', '', str(nota_digitada)).strip().upper()
    if not nota_limpa: return False
    
    meses = {1:'JANEIRO', 2:'FEVEREIRO', 3:'MARÇO', 4:'ABRIL', 5:'MAIO', 6:'JUNHO', 7:'JULHO', 8:'AGOSTO', 9:'SETEMBRO', 10:'OUTUBRO', 11:'NOVEMBRO', 12:'DEZEMBRO'}
    mes_atual = meses[datetime.now().month]
    arquivo_blog = os.path.join(pasta_fretes, f"FRETES_{mes_atual}.xlsx")
    
    if not os.path.exists(arquivo_blog): return False
    
    temp_file_path = os.path.join(tempfile.gettempdir(), f"temp_blog_check_{int(time.time())}.xlsx")
    try:
        shutil.copy2(arquivo_blog, temp_file_path)
        wb = openpyxl.load_workbook(temp_file_path, data_only=True)
        ws = wb.active
        for row in range(3, ws.max_row + 1):
            cell_value = ws.cell(row=row, column=3).value
            if cell_value is not None:
                val_str = str(cell_value).strip().upper()
                val_str = re.sub(r'\.0$', '', val_str)
                if val_str == nota_limpa: return True
    except Exception as e: log.warning("Erro ao verificar nota duplicada: %s", e)
    finally:
        try:
            if os.path.exists(temp_file_path): os.remove(temp_file_path)
        except: pass
    return False