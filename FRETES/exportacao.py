import os
import re
import tempfile
import webbrowser
import sqlite3
import pandas as pd
from datetime import datetime
import html as html_escape_module
import logging
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

def sanitizar_nome_arquivo(nome: str, max_len: int = 80) -> str:
    nome = re.sub(r'[^\w\s\-]', '', nome.strip(), flags=re.UNICODE).replace(' ', '_')
    return re.sub(r'\.\.+', '', nome)[:max_len] if nome else "sem_nome"

def converter_moeda_export(valor_str):
    if pd.isna(valor_str) or valor_str == "": return 0.0
    v = str(valor_str).strip().replace('R$', '').replace(' ', '')
    if ',' in v and '.' in v: v = v.replace('.', '').replace(',', '.')
    elif ',' in v: v = v.replace(',', '.')
    try: return float(v)
    except: return 0.0

def processar_exportacao_carga(dados):
    forn = dados['forn'] or "Sem_Fornecedor"
    forn_limpo = sanitizar_nome_arquivo(forn)
    arquivo_aberto_atual = dados['arquivo_aberto_atual']
    
    d_s = None
    if arquivo_aberto_atual:
        nome_base_antigo = os.path.basename(arquivo_aberto_atual)
        match_data = re.search(r'(\d{2}-\d{2}-\d{4}_\d{2}-\d{2}-\d{2})', nome_base_antigo)
        if match_data:
            d_s = match_data.group(1)
    
    if not d_s:
        d_s = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
        
    num_nota_limpo = re.sub(r'[^\w\-]', '', dados['num_nota'])[:20]
    sufixo_nota = f"_NF_{num_nota_limpo}" if num_nota_limpo else ""
    
    num_pedido_limpo = re.sub(r'[^\w\-]', '', dados['num_pedido'])[:20]
    sufixo_pedido = f"_PED_{num_pedido_limpo}" if num_pedido_limpo else ""
    
    is_pendente = dados['status_pagamento'] == "PENDENTES"
    sufixo_pend = "_PENDENTE" if is_pendente else ""
    
    nome_excel = f"{forn_limpo}_{d_s}{sufixo_nota}{sufixo_pedido}{sufixo_pend}"
    
    if not dados['itens']: 
        return arquivo_aberto_atual
        
    caminho_excel = os.path.join(dados['pasta_arquivo'], f"{nome_excel}.xlsx")
    df_exp = pd.DataFrame(dados['itens'])
    
    with pd.ExcelWriter(caminho_excel, engine='openpyxl') as writer:
        df_exp.to_excel(writer, index=False, sheet_name='Precificacao')
        worksheet = writer.sheets['Precificacao']
        fill_header = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        font_header = Font(color="FFFFFF", bold=True)
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align_center = Alignment(horizontal="center", vertical="center")
        
        for col_num, value in enumerate(df_exp.columns.values):
            cell = worksheet.cell(row=1, column=col_num + 1)
            cell.fill = fill_header; cell.font = font_header; cell.border = border_thin; cell.alignment = align_center

        for col_num, col_name in enumerate(df_exp.columns):
            max_length = len(str(col_name))
            for row_num, val in enumerate(df_exp[col_name]):
                cell = worksheet.cell(row=row_num + 2, column=col_num + 1)
                cell.border = border_thin
                if col_name != "Produto": cell.alignment = align_center
                if val is not None: max_length = max(max_length, len(str(val)))
            worksheet.column_dimensions[get_column_letter(col_num + 1)].width = (max_length + 2)
    
    if arquivo_aberto_atual and "_PENDENTE" in arquivo_aberto_atual:
        if arquivo_aberto_atual != caminho_excel:
            try:
                if os.path.exists(arquivo_aberto_atual):
                    os.remove(arquivo_aberto_atual)
            except Exception as ex:
                log.error(f"Erro ao apagar ficheiro pendente antigo: {ex}")
                
    novo_arquivo_retorno = caminho_excel

    # ---------------- ESPELHO HTML ----------------
    e = html_escape_module.escape
    ignorar_html = ["NOTA", "EMISSAO", "CHEGADA", "TIPO FRETE", "VLR TERCEIRO"]
    
    colunas_html = "".join([f"<th>{e(str(k))}</th>" for k in dados['itens'][0].keys() if k not in ignorar_html])
    linhas_html = ""
    for d in dados['itens']:
        linhas_html += "<tr>" + "".join([
            f"<td class='novo-custo'>{e(str(v))}</td>" if k=="NOVO CUSTO" 
            else f"<td class='venda-prazo'>{e(str(v))}</td>" if k in ["VENDA (R$)", "PRAZO (R$)", "MKP REAL"] 
            else f"<td class='prod-nome'>{e(str(v))}</td>" if k=="Produto" 
            else f"<td>{e(str(v))}</td>" 
            for k,v in d.items() if k not in ignorar_html
        ]) + "</tr>"
    
    forn_safe = e(forn); pedido_safe = e(num_pedido_limpo)
    texto_pedido = f" | <b>Pedido FDC:</b> <span style='color: #c0392b;'>{pedido_safe}</span>" if pedido_safe else ""
    resumo_safe = e(dados['resumo_texto'])
    
    nota_safe = e(dados['num_nota'])
    frete_safe = e(dados['tipo_frete'])
    pag_safe = e(dados['status_pagamento'])

    alerta_pendente_html = ""
    if is_pendente:
        alerta_pendente_html = "<div style='color: #c0392b; border: 2px dashed #c0392b; padding: 15px; margin-bottom: 20px; font-size: 16px; font-weight: bold; text-align: center; background-color: #fadbd8; text-transform: uppercase;'>⚠️ ATENÇÃO: ESTA CARGA ESTÁ PENDENTE DE BOLETOS (AUDITORIA INCOMPLETA) ⚠️</div>"
        
    html = f"""
    <html><head><meta charset="utf-8"><title>Espelho de Carga - {forn_safe}</title><style>
        body {{ font-family: 'Segoe UI', Arial, sans-serif; padding: 20px; color: #333; }}
        h2 {{ color: #2C3E50; margin-bottom: 5px; }} .info {{ font-size: 14px; color: #7F8C8D; margin-bottom: 20px; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 10px; font-size: 12px; }}
        th, td {{ border: 1px solid #BDC3C7; padding: 8px; text-align: center; }} th {{ background-color: #34495E; color: white; }}
        .prod-nome {{ text-align: left; }} .novo-custo {{ font-weight: bold; color: #D35400; background-color: #FDEBD0; }}
        .venda-prazo {{ font-weight: bold; color: #0E6655; background-color: #E8F8F5; }}
        .resumo {{ margin-top: 30px; border: 2px solid #2C3E50; padding: 15px; background: #ECF0F1; font-size: 14px; text-align: center; }}
        @media print {{ .no-print {{ display: none !important; }} }}
    </style></head><body>
    <h2>ESPELHO DE PRECIFICAÇÃO E VENDAS - BRUNO ELETROMÓVEIS</h2>
    {alerta_pendente_html}
    <div class="info">
        <b>Fornecedor:</b> {forn_safe}{texto_pedido} | <b>Nota Fiscal:</b> {nota_safe} | <b>Regime:</b> {e(dados['regime'])} <br>
        <b>Tipo de Frete:</b> {frete_safe} | <b>Pagamento:</b> {pag_safe} | <b>Data:</b> {datetime.now().strftime('%d/%m/%Y às %H:%M')}
    </div>
    <table><tr>{colunas_html}</tr>{linhas_html}</table>
    <div class="resumo"><p><b>DEMONSTRATIVO FINANCEIRO DA COMPRA:</b></p><p style="font-size:16px; color:#c0392b;"><b>{resumo_safe}</b></p></div>
    <div class="no-print" style="margin-top: 40px; text-align: center;">
        <button onclick="window.print()" style="padding: 15px 30px; font-size: 18px; cursor: pointer; background-color: #2980b9; color: white; border: none; border-radius: 5px; font-weight: bold; box-shadow: 0 4px 6px rgba(0,0,0,0.1);">🖨️ ENVIAR PARA IMPRESSORA</button>
    </div></body></html>
    """
    
    p_html = os.path.join(tempfile.gettempdir(), "espelho_impressao_temp.html")
    with open(p_html, 'w', encoding='utf-8') as f: f.write(html)
    webbrowser.open('file://' + os.path.realpath(p_html))

    # ---------------- B-LOG ----------------
    if not is_pendente:
        _atualizar_planilha_blog(dados)

    return novo_arquivo_retorno

def _atualizar_planilha_blog(dados):
    from tkinter import messagebox # Import necessário para o pop-up de erro

    forn = dados['forn']
    if not forn: return
    
    uf_fornecedor = ""
    try:
        with sqlite3.connect(dados['db_path'], timeout=10) as conn:
            res = conn.cursor().execute("SELECT uf FROM fornecedores WHERE fabricante=?", (forn,)).fetchone()
            if res: uf_fornecedor = res[0]
    except: pass
            
    val_nota = dados['total_mercadoria_compra'] + dados['total_ipi_compra']
    val_frete_linhas = dados['total_frete_compra']
    
    tipo_frete = dados['tipo_frete']
    val_terceiro = converter_moeda_export(dados['val_terceiro_str']) if tipo_frete == "TERCEIRIZADO" else 0.0
    
    frete_combinado_perc = val_frete_linhas / val_nota if val_nota > 0 else 0.0
        
    if tipo_frete == "CIF":
        val_blog = 0.0; receita = 0.0
    else: 
        val_blog = val_frete_linhas; receita = val_blog - val_terceiro
        
    meses = {1:'JANEIRO', 2:'FEVEREIRO', 3:'MARÇO', 4:'ABRIL', 5:'MAIO', 6:'JUNHO', 7:'JULHO', 8:'AGOSTO', 9:'SETEMBRO', 10:'OUTUBRO', 11:'NOVEMBRO', 12:'DEZEMBRO'}
    mes_atual = meses[datetime.now().month]
    arquivo_blog = os.path.join(dados['pasta_fretes'], f"FRETES_{mes_atual}.xlsx")
    novo_arquivo = not os.path.exists(arquivo_blog)
    
    if novo_arquivo:
        wb = Workbook()
        ws = wb.active
        ws.title = "FRETES"
        ws.merge_cells('A1:D1')
        ws['A1'] = "CONTROLE DE FRETE BLOG"
        ws['A1'].font = Font(bold=True)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        headers = {1: "DATA DE EMISSÃO", 2: "DATA DE CHEGADA", 3: "NÚMERO DA NOTA", 4: "FABRICANTE", 5: "UF", 6: "VALOR NOTA", 7: "TIPO FRETE", 8: "FRETE COMBINADO", 10: "VALOR B-LOG", 11: "$ VENDIDO ou TERCEIRIZADO", 12: "RECEITA"}
        for col, h in headers.items():
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin = Side(border_style="thin", color="000000")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws.column_dimensions[get_column_letter(col)].width = 18
    else:
        wb = openpyxl.load_workbook(arquivo_blog)
        ws = wb.active
        
    nota_procurada = re.sub(r'\.0$', '', dados['num_nota']).upper()
    row_idx = 3
    
    while ws.cell(row=row_idx, column=1).value or ws.cell(row=row_idx, column=3).value or ws.cell(row=row_idx, column=4).value:
        cel_nota = ws.cell(row=row_idx, column=3).value
        if cel_nota:
            nota_planilha = re.sub(r'\.0$', '', str(cel_nota).strip()).upper()
            if nota_planilha == nota_procurada:
                break
        row_idx += 1
        
    ws.cell(row=row_idx, column=1, value=dados['dt_emissao'])
    ws.cell(row=row_idx, column=2, value=dados['dt_chegada'])
    ws.cell(row=row_idx, column=3, value=dados['num_nota'])
    ws.cell(row=row_idx, column=4, value=dados['forn'])
    ws.cell(row=row_idx, column=5, value=uf_fornecedor)
    ws.cell(row=row_idx, column=6, value=val_nota).number_format = 'R$ #,##0.00'
    ws.cell(row=row_idx, column=7, value=tipo_frete)
    ws.cell(row=row_idx, column=8, value=frete_combinado_perc).number_format = '0.0%'
    
    ws.cell(row=row_idx, column=10, value=val_blog).number_format = 'R$ #,##0.00'
    ws.cell(row=row_idx, column=11, value=val_terceiro if val_terceiro > 0 else 0).number_format = 'R$ #,##0.00'
    ws.cell(row=row_idx, column=12, value=receita).number_format = 'R$ #,##0.00'
    
    fill_azul = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    fill_laranja = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
    
    ws.cell(row=row_idx, column=11).fill = fill_laranja
    ws.cell(row=row_idx, column=12).fill = fill_azul
    
    # === AQUI ESTÁ O BLOQUEIO DA TELA PRETA (PermissionError) ===
    try:
        wb.save(arquivo_blog)
    except PermissionError:
        messagebox.showerror(
            "Aviso de Segurança", 
            f"A planilha de fretes ({os.path.basename(arquivo_blog)}) está aberta no Excel!\n\nPor favor, feche o Excel e audite a nota novamente para guardar os dados da B-LOG."
        )

def salvar_edicao_precos(dados):
    """
    Atualiza SOMENTE VENDA (R$), PRAZO (R$) e MKP REAL no xlsx existente.
    - Salva no MESMO arquivo (sem criar novo)
    - NÃO toca na planilha BLOG
    - Regenera o espelho HTML
    """
    caminho = dados.get('arquivo_aberto_atual')
    if not caminho or not os.path.exists(caminho):
        return False, "Nenhum arquivo aberto. Abra uma precificação salva antes de editar."

    # Índice de preços por código do produto
    precos = {}
    for item in dados['itens']:
        cod = str(item.get('Código', '')).strip()
        if cod:
            precos[cod] = {
                'VENDA (R$)': item.get('VENDA (R$)', ''),
                'PRAZO (R$)': item.get('PRAZO (R$)', ''),
                'MKP REAL':   item.get('MKP REAL', ''),
            }

    # Abrir xlsx e localizar colunas
    try:
        wb = openpyxl.load_workbook(caminho)
    except PermissionError:
        return False, f"O arquivo está aberto no Excel!\nFeche-o e tente novamente.\n{os.path.basename(caminho)}"

    ws = wb['Precificacao']
    header = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    col_cod   = header.get('Código')
    col_venda = header.get('VENDA (R$)')
    col_prazo = header.get('PRAZO (R$)')
    col_mkp   = header.get('MKP REAL')

    if not all([col_cod, col_venda, col_prazo]):
        return False, "Colunas de preço não encontradas no arquivo. Estrutura incompatível."

    # Atualizar somente as células de preço
    for row in range(2, ws.max_row + 1):
        cod = str(ws.cell(row=row, column=col_cod).value or '').strip()
        if cod in precos:
            ws.cell(row=row, column=col_venda).value = precos[cod]['VENDA (R$)']
            ws.cell(row=row, column=col_prazo).value = precos[cod]['PRAZO (R$)']
            if col_mkp:
                ws.cell(row=row, column=col_mkp).value = precos[cod]['MKP REAL']

    try:
        wb.save(caminho)
    except PermissionError:
        return False, f"Não foi possível salvar — o arquivo está aberto no Excel!\n{os.path.basename(caminho)}"

    # Regenerar espelho HTML (sem BLOG)
    _gerar_espelho_html(dados)

    return True, caminho


def _gerar_espelho_html(dados):
    """Gera e abre o espelho HTML sem alterar o BLOG."""
    import html as html_escape_module
    e = html_escape_module.escape

    forn = dados['forn'] or "Sem_Fornecedor"
    num_pedido_limpo = re.sub(r'[^\w\-]', '', dados['num_pedido'])[:20]
    is_pendente = dados['status_pagamento'] == "PENDENTES"

    ignorar_html = ["NOTA", "EMISSAO", "CHEGADA", "TIPO FRETE", "VLR TERCEIRO"]
    colunas_html = "".join([f"<th>{e(str(k))}</th>" for k in dados['itens'][0].keys() if k not in ignorar_html])
    linhas_html = ""
    for d in dados['itens']:
        linhas_html += "<tr>" + "".join([
            f"<td class='novo-custo'>{e(str(v))}</td>" if k == "NOVO CUSTO"
            else f"<td class='venda-prazo'>{e(str(v))}</td>" if k in ["VENDA (R$)", "PRAZO (R$)", "MKP REAL"]
            else f"<td class='prod-nome'>{e(str(v))}</td>" if k == "Produto"
            else f"<td>{e(str(v))}</td>"
            for k, v in d.items() if k not in ignorar_html
        ]) + "</tr>"

    forn_safe    = e(forn)
    pedido_safe  = e(num_pedido_limpo)
    nota_safe    = e(dados['num_nota'])
    frete_safe   = e(dados['tipo_frete'])
    pag_safe     = e(dados['status_pagamento'])
    resumo_safe  = e(dados['resumo_texto'])
    texto_pedido = f" | <b>Pedido FDC:</b> <span style='color:#c0392b;'>{pedido_safe}</span>" if pedido_safe else ""

    alerta_pendente_html = ""
    if is_pendente:
        alerta_pendente_html = "<div style='color:#c0392b;border:2px dashed #c0392b;padding:15px;margin-bottom:20px;font-size:16px;font-weight:bold;text-align:center;background-color:#fadbd8;text-transform:uppercase;'>⚠️ ATENÇÃO: ESTA CARGA ESTÁ PENDENTE DE BOLETOS (AUDITORIA INCOMPLETA) ⚠️</div>"

    html = f"""
    <html><head><meta charset="utf-8"><title>Espelho de Carga - {forn_safe}</title><style>
        body{{font-family:'Segoe UI',Arial,sans-serif;padding:20px;color:#333;}}
        h2{{color:#2C3E50;margin-bottom:5px;}}.info{{font-size:14px;color:#7F8C8D;margin-bottom:20px;}}
        table{{width:100%;border-collapse:collapse;margin-top:10px;font-size:12px;}}
        th,td{{border:1px solid #BDC3C7;padding:8px;text-align:center;}}th{{background-color:#34495E;color:white;}}
        .prod-nome{{text-align:left;}}.novo-custo{{font-weight:bold;color:#D35400;background-color:#FDEBD0;}}
        .venda-prazo{{font-weight:bold;color:#0E6655;background-color:#E8F8F5;}}
        .resumo{{margin-top:30px;border:2px solid #2C3E50;padding:15px;background:#ECF0F1;font-size:14px;text-align:center;}}
        @media print{{.no-print{{display:none!important;}}}}
    </style></head><body>
    <h2>ESPELHO DE PRECIFICAÇÃO E VENDAS — EDIÇÃO DE PREÇOS</h2>
    {alerta_pendente_html}
    <div class="info">
        <b>Fornecedor:</b> {forn_safe}{texto_pedido} | <b>Nota Fiscal:</b> {nota_safe} | <b>Regime:</b> {e(dados['regime'])} <br>
        <b>Tipo de Frete:</b> {frete_safe} | <b>Pagamento:</b> {pag_safe} | <b>Atualizado em:</b> {datetime.now().strftime('%d/%m/%Y às %H:%M')}
    </div>
    <table><tr>{colunas_html}</tr>{linhas_html}</table>
    <div class="resumo"><p><b>DEMONSTRATIVO FINANCEIRO DA COMPRA:</b></p>
    <p style="font-size:16px;color:#c0392b;"><b>{resumo_safe}</b></p></div>
    <div class="no-print" style="margin-top:40px;text-align:center;">
        <button onclick="window.print()" style="padding:15px 30px;font-size:18px;cursor:pointer;background-color:#2980b9;color:white;border:none;border-radius:5px;font-weight:bold;">🖨️ ENVIAR PARA IMPRESSORA</button>
    </div></body></html>"""

    import tempfile, webbrowser
    p_html = os.path.join(tempfile.gettempdir(), "espelho_impressao_temp.html")
    with open(p_html, 'w', encoding='utf-8') as f:
        f.write(html)
    webbrowser.open('file://' + os.path.realpath(p_html))