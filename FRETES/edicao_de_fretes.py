import os
import glob
import tkinter as tk
from tkinter import ttk, messagebox
import ttkbootstrap as ttkb
import openpyxl
import tempfile
import webbrowser
from datetime import datetime

def abrir_modulo_fretes(root, pasta_fretes):
    """
    Abre a interface de gerenciamento e edição das planilhas de frete.
    """
    # TRAVA DE SILÊNCIO: Desliga validações da tela principal
    if hasattr(root, 'ignorando_validacao'):
        root.ignorando_validacao = True

    arquivos = glob.glob(os.path.join(pasta_fretes, "FRETES_*.xlsx"))
    if not arquivos:
        messagebox.showinfo("Aviso", "Nenhuma planilha de frete encontrada na pasta FRETES.")
        if hasattr(root, 'ignorando_validacao'):
            root.ignorando_validacao = False
        return

    arquivos.sort(key=os.path.getmtime)
    
    mapa_arquivos = {}
    for arq in arquivos:
        nome_arquivo = os.path.basename(arq)
        mes = nome_arquivo.replace("FRETES_", "").replace(".xlsx", "")
        mapa_arquivos[mes] = arq

    janela = tk.Toplevel(root)
    janela.title("Gerenciador de Fretes B-LOG")
    
    # AJUSTE DE TELA INTELIGENTE: Proporcional a qualquer monitor
    screen_w = root.winfo_screenwidth()
    screen_h = root.winfo_screenheight()
    
    w = int(screen_w * 0.90) 
    h = int(screen_h * 0.82) 
    
    x = int((screen_w - w) / 2) 
    y = int((screen_h - h) / 2) - 20 
    
    janela.geometry(f"{w}x{h}+{x}+{y}")
    janela.transient(root) 
    janela.grab_set()      

    arquivo_atual_path = None

    def fechar_janela():
        if hasattr(root, 'ignorando_validacao'):
            root.ignorando_validacao = False
        janela.destroy()

    janela.protocol("WM_DELETE_WINDOW", fechar_janela)
    
    # --- BARRA SUPERIOR (SELEÇÃO DE MÊS) ---
    f_topo = ttkb.Frame(janela, bootstyle="secondary", padding=15)
    f_topo.pack(fill="x", side="top")

    ttkb.Label(f_topo, text="📅 SELECIONE O MÊS:", font=("Segoe UI", 12, "bold"), bootstyle="inverse-secondary").pack(side="left", padx=10)
    
    combo_mes = ttk.Combobox(f_topo, values=list(mapa_arquivos.keys()), state="readonly", font=("Segoe UI", 12), width=20)
    combo_mes.pack(side="left", padx=10)

    lbl_dica = tk.Label(f_topo, text="💡 DICA: Dê um DUPLO CLIQUE em qualquer valor da tabela para editá-lo.", font=("Segoe UI", 10, "italic", "bold"), bg="#2C3E50", fg="#F1C40F")
    lbl_dica.pack(side="right", padx=20)

    # --- NOVO PAINEL DE TOTAIS FIXO NO TOPO ---
    f_totais = ttkb.Frame(janela, bootstyle="info", padding=10)
    f_totais.pack(fill="x", padx=10, pady=(10, 0))
    
    var_resumo_totais = tk.StringVar(value="TOTAIS DO MÊS: Selecione uma planilha...")
    lbl_totais_display = tk.Label(f_totais, textvariable=var_resumo_totais, font=("Segoe UI", 13, "bold"), bg="#5bc0de", fg="white")
    lbl_totais_display.pack()

    # --- ÁREA DA TABELA ---
    f_grid = ttkb.Frame(janela, padding=10)
    f_grid.pack(fill="both", expand=True)

    style = ttk.Style(janela)
    style.configure("Treeview.Heading", font=("Segoe UI", 10, "bold"))
    style.configure("Treeview", font=("Segoe UI", 10), rowheight=35)

    scroll_y = ttk.Scrollbar(f_grid, orient="vertical")
    scroll_y.pack(side="right", fill="y")
    scroll_x = ttk.Scrollbar(f_grid, orient="horizontal")
    scroll_x.pack(side="bottom", fill="x")

    img_espacadora = tk.PhotoImage(width=1, height=35)
    janela.img_espacadora = img_espacadora 

    tree = ttk.Treeview(f_grid, yscrollcommand=scroll_y.set, xscrollcommand=scroll_x.set, selectmode="browse")
    tree["show"] = "tree headings"
    tree.column("#0", width=1, minwidth=1, stretch=tk.NO) 

    scroll_y.config(command=tree.yview)
    scroll_x.config(command=tree.xview)
    tree.pack(fill="both", expand=True)

    tree.tag_configure('par', background='#F8F9F9', foreground='black')
    tree.tag_configure('impar', background='#EBF5FB', foreground='black')

    # --- FUNÇÃO PARA CALCULAR OS TOTAIS NA TELA ---
    def atualizar_totais():
        if not arquivo_atual_path: return
        
        colunas = tree["columns"]
        linhas = [tree.item(item, "values") for item in tree.get_children()]
        
        tot_nota = 0.0
        tot_blog = 0.0
        tot_terceiro = 0.0
        tot_receita = 0.0
        
        for row in linhas:
            for i, val in enumerate(row):
                if i >= len(colunas): continue
                col_name = str(colunas[i]).upper()
                
                try:
                    v_str = str(val).replace("R$", "").replace("%", "").strip()
                    if v_str:
                        if "," in v_str and "." in v_str:
                            v_str = v_str.replace(".", "").replace(",", ".")
                        elif "," in v_str:
                            v_str = v_str.replace(",", ".")
                        val_float = float(v_str)
                        
                        if "VALOR NOTA" in col_name or i == 5:
                            tot_nota += val_float
                        elif "VALOR B-LOG" in col_name or i == 9:
                            tot_blog += val_float
                        elif "TERCEIRIZADO" in col_name or "$" in col_name or i == 10:
                            tot_terceiro += val_float
                        elif "RECEITA" in col_name or i == 11:
                            tot_receita += val_float
                except: pass
                
        def formatar(v):
            return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
            
        resumo = f"🛒 VALOR NOTA: {formatar(tot_nota)}    |    🚚 VALOR B-LOG: {formatar(tot_blog)}    |    💸 TERCEIROS: {formatar(tot_terceiro)}    |    ✅ RECEITA LÍQUIDA: {formatar(tot_receita)}"
        var_resumo_totais.set(resumo)

    def carregar_planilha(event=None):
        nonlocal arquivo_atual_path
        mes_selecionado = combo_mes.get()
        if not mes_selecionado: return

        arquivo_atual_path = mapa_arquivos[mes_selecionado]
        tree.delete(*tree.get_children())
        tree["columns"] = ()

        try:
            wb = openpyxl.load_workbook(arquivo_atual_path, data_only=True)
            ws = wb.active

            colunas = []
            max_col = ws.max_column
            for col in range(1, max_col + 1):
                valor = ws.cell(row=2, column=col).value
                colunas.append(str(valor) if valor else f"Col {col}")

            tree["columns"] = colunas

            for col in colunas:
                tree.heading(col, text=col)
                tree.column(col, width=150, anchor="center")

            linhas_carregadas = []
            for row_idx in range(3, ws.max_row + 1):
                linha_dados = []
                vazia = True
                for col_idx in range(1, max_col + 1):
                    val = ws.cell(row=row_idx, column=col_idx).value
                    if val is not None: vazia = False
                    linha_dados.append(val)
                if not vazia:
                    linhas_carregadas.append(linha_dados)

            def get_data_chegada_load(linha):
                try:
                    data_str = str(linha[1]).strip()
                    if "/" in data_str:
                        p = data_str.split("/")
                        if len(p) == 3: return datetime(int(p[2]), int(p[1]), int(p[0]))
                except: pass
                return datetime.max
            
            linhas_carregadas.sort(key=get_data_chegada_load)

            for row_idx, row_vals in enumerate(linhas_carregadas):
                linha_formatada = []
                for col_idx, val in enumerate(row_vals, start=1):
                    texto = str(val) if val is not None else ""
                    if isinstance(val, float) or isinstance(val, int):
                        nome_col = colunas[col_idx-1].upper() if col_idx <= len(colunas) else ""
                        if "VALOR" in nome_col or "RECEITA" in nome_col or "$" in nome_col or col_idx in [6, 9, 10, 11, 12]:
                            texto = f"R$ {float(val):,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                        elif "COMBINADO" in nome_col or "%" in nome_col or col_idx == 8:
                            texto = f"{float(val)*100:.2f}%"
                    linha_formatada.append(texto)
                
                tag = 'par' if row_idx % 2 == 0 else 'impar'
                tree.insert("", "end", image=img_espacadora, values=linha_formatada, tags=(tag,))
                
            # Atualiza os totais logo após carregar os dados na tabela
            atualizar_totais()

        except Exception as e:
            messagebox.showerror("Erro", f"Erro ao carregar a planilha:\n{e}")

    combo_mes.bind("<<ComboboxSelected>>", carregar_planilha)

    if mapa_arquivos:
        def abrir_lista_sozinha():
            combo_mes.focus_force()
            try:
                janela.tk.call('ttk::combobox::Post', combo_mes)
            except tk.TclError: pass
        janela.after(500, abrir_lista_sozinha)

    # --- LÓGICA DE EDIÇÃO ---
    def on_double_click(event):
        if not arquivo_atual_path: return
        item = tree.identify_row(event.y)
        column = tree.identify_column(event.x)
        if not item or not column or column == "#0": return 

        col_idx = int(column[1:]) - 1
        valor_atual = tree.item(item, "values")[col_idx]

        popup = tk.Toplevel(janela)
        popup.wm_overrideredirect(True)
        popup.attributes("-topmost", True) 

        x_root = tree.winfo_rootx() + event.x
        y_root = tree.winfo_rooty() + event.y
        popup.geometry(f"200x65+{x_root-100}+{y_root-30}")

        frame_popup = tk.Frame(popup, bg="#2980B9", bd=2, relief="solid")
        frame_popup.pack(fill="both", expand=True)

        lbl = tk.Label(frame_popup, text="EDITAR VALOR (ENTER salva):", bg="#2980B9", fg="white", font=("Segoe UI", 8, "bold"))
        lbl.pack(pady=(2, 0))

        entry_edit = tk.Entry(frame_popup, font=("Segoe UI", 12, "bold"), justify="center", bg="#FFF8E1", fg="#c0392b")
        entry_edit.pack(padx=5, pady=2, fill="x")
        
        valor_limpo = valor_atual.replace("R$", "").replace("%", "").strip()
        entry_edit.insert(0, valor_limpo)
        entry_edit.select_range(0, tk.END)
        entry_edit.focus_force()

        def salvar_edicao(e=None):
            try:
                novo_valor = entry_edit.get().strip()
                valores = list(tree.item(item, "values"))
                valores[col_idx] = novo_valor
                tree.item(item, values=valores)
            except: pass
            popup.destroy()
            # Atualiza o painel de totais sempre que um valor for editado!
            atualizar_totais()

        def cancelar_edicao(e=None):
            popup.destroy()

        entry_edit.bind("<Return>", salvar_edicao)
        entry_edit.bind("<Escape>", cancelar_edicao)
        popup.bind("<FocusOut>", lambda e: popup.after(100, salvar_edicao))

    tree.bind("<Double-1>", on_double_click)

    f_rodape = ttkb.Frame(janela, padding=15)
    f_rodape.pack(fill="x", side="bottom")

    def imprimir_relatorio():
        if not arquivo_atual_path:
            messagebox.showwarning("Aviso", "Nenhuma planilha carregada para imprimir.")
            return

        colunas = tree["columns"]
        linhas = [tree.item(item, "values") for item in tree.get_children()]
        if not linhas: return

        totais = [0.0] * len(colunas)
        for row in linhas:
            for i, val in enumerate(row):
                try:
                    v_str = str(val).replace("R$", "").replace("%", "").strip()
                    if "," in v_str and "." in v_str:
                        v_str = v_str.replace(".", "").replace(",", ".")
                    elif "," in v_str:
                        v_str = v_str.replace(",", ".")
                    totais[i] += float(v_str)
                except: pass

        html = f"""
        <html><head><meta charset="utf-8"><title>Relatório de Fretes - {combo_mes.get()}</title><style>
            body {{ font-family: 'Segoe UI', Arial, sans-serif; padding: 20px; color: #333; }}
            h2 {{ color: #2C3E50; text-align: center; margin-bottom: 20px; text-transform: uppercase; }}
            table {{ width: 100%; border-collapse: collapse; font-size: 13px; }}
            th, td {{ border: 1px solid #BDC3C7; padding: 8px; text-align: center; }}
            th {{ background-color: #34495E; color: white; }}
            .par {{ background-color: #F8F9F9; color: black; }} 
            .impar {{ background-color: #EBF5FB; color: black; }}
            .totais {{ background-color: #2C3E50; color: white; font-weight: bold; font-size: 14px; }}
            @media print {{ .no-print {{ display: none !important; }} }}
        </style></head><body>
        <h2>📊 RELATÓRIO MENSAL DE FRETES - {combo_mes.get()}</h2>
        <table><thead><tr>
        """
        for col in colunas: html += f"<th>{col}</th>"
        html += "</tr></thead><tbody>"

        for idx, row in enumerate(linhas):
            classe = "par" if idx % 2 == 0 else "impar"
            html += f"<tr class='{classe}'>"
            for val in row: html += f"<td>{val}</td>"
            html += "</tr>"

        html += "<tr class='totais'><td>TOTAIS</td>"
        for i in range(1, len(colunas)):
            nome_col = colunas[i].upper()
            if "VALOR" in nome_col or "RECEITA" in nome_col or "$" in nome_col or i in [5, 8, 9, 10, 11]:
                html += f"<td>R$ {totais[i]:,.2f}</td>".replace(",", "X").replace(".", ",").replace("X", ".")
            else:
                html += "<td>-</td>"
        html += "</tr></tbody></table>"

        html += """
        <div class="no-print" style="margin-top: 30px; text-align: center;">
            <button onclick="window.print()" style="padding: 12px 25px; font-size: 16px; cursor: pointer; background-color: #2980b9; color: white; border: none; border-radius: 5px; font-weight: bold;">🖨️ ENVIAR PARA IMPRESSORA</button>
        </div></body></html>
        """

        p_html = os.path.join(tempfile.gettempdir(), "Relatorio_Fretes_Impressao.html")
        with open(p_html, 'w', encoding='utf-8') as f: f.write(html)
        webbrowser.open('file://' + os.path.realpath(p_html))

    def salvar_planilha():
        if not arquivo_atual_path:
            messagebox.showwarning("Aviso", "Nenhuma planilha carregada.")
            return

        try:
            wb = openpyxl.load_workbook(arquivo_atual_path)
            ws = wb.active

            linhas_dados = []
            for item in tree.get_children():
                linhas_dados.append(list(tree.item(item, "values")))

            def get_data_chegada(linha):
                try:
                    data_str = str(linha[1]).strip()
                    if "/" in data_str:
                        partes = data_str.split("/")
                        if len(partes) == 3: return datetime(int(partes[2]), int(partes[1]), int(partes[0]))
                except: pass
                return datetime.max 

            linhas_dados.sort(key=get_data_chegada)

            max_row = ws.max_row
            if max_row >= 3:
                ws.delete_rows(3, max_row - 2)

            colunas = [str(ws.cell(row=2, column=col).value).upper() for col in range(1, ws.max_column + 1)]

            for i, linha in enumerate(linhas_dados, start=3):
                for j, val in enumerate(linha, start=1):
                    texto = str(val).strip()
                    nome_coluna = colunas[j-1] if j <= len(colunas) else ""
                    
                    v_limpo = texto.replace("R$", "").replace("%", "").strip()
                    num_val = None
                    is_numeric = False
                    
                    if v_limpo != "":
                        try:
                            v_conv = v_limpo
                            if "," in v_conv and "." in v_conv: v_conv = v_conv.replace(".", "").replace(",", ".")
                            elif "," in v_conv: v_conv = v_conv.replace(",", ".")
                            num_val = float(v_conv)
                            is_numeric = True
                        except: pass

                    celula = ws.cell(row=i, column=j)
                    
                    if is_numeric:
                        if "VALOR" in nome_coluna or "RECEITA" in nome_coluna or "$" in nome_coluna or j in [6, 9, 10, 11, 12]:
                            celula.value = num_val
                            celula.number_format = 'R$ #,##0.00'
                        elif "COMBINADO" in nome_coluna or "%" in nome_coluna or j == 8:
                            celula.value = (num_val / 100.0) if num_val > 1.0 else num_val
                            celula.number_format = '0.0%'
                        else:
                            celula.value = num_val
                    else:
                        celula.value = texto if texto != "" else None
            
            wb.save(arquivo_atual_path)

            tree.delete(*tree.get_children())
            for row_idx, linha in enumerate(linhas_dados, start=3):
                linha_formatada = []
                for j, val in enumerate(linha, start=1):
                    texto = str(val).strip()
                    nome_coluna = colunas[j-1] if j <= len(colunas) else ""
                    
                    v_limpo = texto.replace("R$", "").replace("%", "").strip()
                    if v_limpo != "":
                        try:
                            v_conv = v_limpo
                            if "," in v_conv and "." in v_conv: v_conv = v_conv.replace(".", "").replace(",", ".")
                            elif "," in v_conv: v_conv = v_conv.replace(",", ".")
                            num_f = float(v_conv)
                            
                            if "VALOR" in nome_coluna or "RECEITA" in nome_coluna or "$" in nome_coluna or j in [6, 9, 10, 11, 12]:
                                texto = f"R$ {num_f:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
                            elif "COMBINADO" in nome_coluna or "%" in nome_coluna or j == 8:
                                texto = f"{num_f if num_f > 1.0 else num_f*100:,.2f}%".replace(".", ",")
                        except: pass
                    linha_formatada.append(texto)
                
                tag = 'par' if row_idx % 2 == 0 else 'impar'
                tree.insert("", "end", image=img_espacadora, values=linha_formatada, tags=(tag,))

            messagebox.showinfo("Sucesso", "Alterações salvas com sucesso!\nA planilha foi ordenada pela DATA DE CHEGADA.")
        except PermissionError:
            messagebox.showerror("Erro", "O arquivo está aberto no Excel. Feche-o e tente novamente.")
        except Exception as e:
            messagebox.showerror("Erro", f"Ocorreu um erro ao salvar:\n{e}")

    btn_imprimir = ttkb.Button(f_rodape, text="🖨️ IMPRIMIR PLANILHA", bootstyle="info", command=imprimir_relatorio)
    btn_imprimir.pack(side="left", padx=10)

    btn_fechar = ttkb.Button(f_rodape, text="✖ FECHAR", bootstyle="danger", command=fechar_janela)
    btn_fechar.pack(side="right", padx=10)

    btn_salvar = ttkb.Button(f_rodape, text="💾 SALVAR ALTERAÇÕES", bootstyle="success", command=salvar_planilha)
    btn_salvar.pack(side="right", padx=10)