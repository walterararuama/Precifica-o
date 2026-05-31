import os
import re
import tempfile
import webbrowser
import sqlite3
import pandas as pd
from datetime import datetime
import logging
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

log = logging.getLogger(__name__)

def sanitizar_nome_arquivo(nome: str, max_len: int = 80) -> str:
    nome = re.sub(r'[^\w\s\-]', '', nome.strip(), flags=re.UNICODE).replace(' ', '_')
    return re.sub(r'\.\.+', '', nome)[:max_len] if nome else "sem_nome"

def converter_moeda_export(valor_str):
    if pd.isna(valor_str) or valor_str == "": return 0.0
    v = str(valor_str).strip().replace('R$', '').replace(' ', '')
    if ',' in v and '.' in v: v = v.replace('.', '').replace(',', '.')
    elif ',' in v: v = v.replace(',', '.')
    try: return float(v)
    except: return 0.0

def processar_exportacao_carga(dados):
    forn = dados['forn'] or "Sem_Fornecedor"
    forn_limpo = sanitizar_nome_arquivo(forn)
    arquivo_aberto_atual = dados['arquivo_aberto_atual']
    
    d_s = None
    if arquivo_aberto_atual:
        nome_base_antigo = os.path.basename(arquivo_aberto_atual)
        match_data = re.search(r'(\d{2}-\d{2}-\d{4}_\d{2}-\d{2}-\d{2})', nome_base_antigo)
        if match_data:
            d_s = match_data.group(1)
    
    if not d_s:
        d_s = datetime.now().strftime("%d-%m-%Y_%H-%M-%S")
        
    num_nota_limpo = re.sub(r'[^\w\-]', '', dados['num_nota'])[:20]
    sufixo_nota = f"_NF_{num_nota_limpo}" if num_nota_limpo else ""
    
    num_pedido_limpo = re.sub(r'[^\w\-]', '', dados['num_pedido'])[:20]
    sufixo_pedido = f"_PED_{num_pedido_limpo}" if num_pedido_limpo else ""
    
    is_pendente = dados['status_pagamento'] == "PENDENTES"
    sufixo_pend = "_PENDENTE" if is_pendente else ""
    
    nome_excel = f"{forn_limpo}_{d_s}{sufixo_nota}{sufixo_pedido}{sufixo_pend}"
    
    if not dados['itens']: 
        return arquivo_aberto_atual
        
    caminho_excel = os.path.join(dados['pasta_arquivo'], f"{nome_excel}.xlsx")
    df_exp = pd.DataFrame(dados['itens'])
    
    with pd.ExcelWriter(caminho_excel, engine='openpyxl') as writer:
        df_exp.to_excel(writer, index=False, sheet_name='Precificacao')
        worksheet = writer.sheets['Precificacao']
        fill_header = PatternFill(start_color="34495E", end_color="34495E", fill_type="solid")
        font_header = Font(color="FFFFFF", bold=True)
        border_thin = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        align_center = Alignment(horizontal="center", vertical="center")
        
        for col_num, value in enumerate(df_exp.columns.values):
            cell = worksheet.cell(row=1, column=col_num + 1)
            cell.fill = fill_header; cell.font = font_header; cell.border = border_thin; cell.alignment = align_center

        for col_num, col_name in enumerate(df_exp.columns):
            max_length = len(str(col_name))
            for row_num, val in enumerate(df_exp[col_name]):
                cell = worksheet.cell(row=row_num + 2, column=col_num + 1)
                cell.border = border_thin
                if col_name != "Produto": cell.alignment = align_center
                if val is not None: max_length = max(max_length, len(str(val)))
            worksheet.column_dimensions[get_column_letter(col_num + 1)].width = (max_length + 2)
    
    if arquivo_aberto_atual and "_PENDENTE" in arquivo_aberto_atual:
        if arquivo_aberto_atual != caminho_excel:
            try:
                if os.path.exists(arquivo_aberto_atual):
                    os.remove(arquivo_aberto_atual)
            except Exception as ex:
                log.error(f"Erro ao apagar ficheiro pendente antigo: {ex}")
                
    novo_arquivo_retorno = caminho_excel

    # ---------------- ESPELHO HTML ----------------
    _gerar_espelho_html(dados)

    # ---------------- B-LOG ----------------
    if not is_pendente:
        _atualizar_planilha_blog(dados)

    return novo_arquivo_retorno

def _atualizar_planilha_blog(dados):
    from tkinter import messagebox # Import necessário para o pop-up de erro

    forn = dados['forn']
    if not forn: return
    
    uf_fornecedor = ""
    try:
        with sqlite3.connect(dados['db_path'], timeout=10) as conn:
            res = conn.cursor().execute("SELECT uf FROM fornecedores WHERE fabricante=?", (forn,)).fetchone()
            if res: uf_fornecedor = res[0]
    except: pass
            
    val_nota = dados['total_mercadoria_compra'] + dados['total_ipi_compra']
    val_frete_linhas = dados['total_frete_compra']
    
    tipo_frete = dados['tipo_frete']
    val_terceiro = converter_moeda_export(dados['val_terceiro_str']) if tipo_frete == "TERCEIRIZADO" else 0.0
    
    frete_combinado_perc = val_frete_linhas / val_nota if val_nota > 0 else 0.0
        
    if tipo_frete == "CIF":
        val_blog = 0.0; receita = 0.0
    else: 
        val_blog = val_frete_linhas; receita = val_blog - val_terceiro
        
    meses = {1:'JANEIRO', 2:'FEVEREIRO', 3:'MARÇO', 4:'ABRIL', 5:'MAIO', 6:'JUNHO', 7:'JULHO', 8:'AGOSTO', 9:'SETEMBRO', 10:'OUTUBRO', 11:'NOVEMBRO', 12:'DEZEMBRO'}
    mes_atual = meses[datetime.now().month]
    arquivo_blog = os.path.join(dados['pasta_fretes'], f"FRETES_{mes_atual}.xlsx")
    novo_arquivo = not os.path.exists(arquivo_blog)
    
    if novo_arquivo:
        wb = Workbook()
        ws = wb.active
        ws.title = "FRETES"
        ws.merge_cells('A1:D1')
        ws['A1'] = "CONTROLE DE FRETE BLOG"
        ws['A1'].font = Font(bold=True)
        ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
        
        headers = {1: "DATA DE EMISSÃO", 2: "DATA DE CHEGADA", 3: "NÚMERO DA NOTA", 4: "FABRICANTE", 5: "UF", 6: "VALOR NOTA", 7: "TIPO FRETE", 8: "FRETE COMBINADO", 10: "VALOR B-LOG", 11: "$ VENDIDO ou TERCEIRIZADO", 12: "RECEITA"}
        for col, h in headers.items():
            cell = ws.cell(row=2, column=col, value=h)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="808080", end_color="808080", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin = Side(border_style="thin", color="000000")
            cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)
            ws.column_dimensions[get_column_letter(col)].width = 18
    else:
        wb = openpyxl.load_workbook(arquivo_blog)
        ws = wb.active
        
    nota_procurada = re.sub(r'\.0$', '', dados['num_nota']).upper()
    row_idx = 3
    
    while ws.cell(row=row_idx, column=1).value or ws.cell(row=row_idx, column=3).value or ws.cell(row=row_idx, column=4).value:
        cel_nota = ws.cell(row=row_idx, column=3).value
        if cel_nota:
            nota_planilha = re.sub(r'\.0$', '', str(cel_nota).strip()).upper()
            if nota_planilha == nota_procurada:
                break
        row_idx += 1
        
    ws.cell(row=row_idx, column=1, value=dados['dt_emissao'])
    ws.cell(row=row_idx, column=2, value=dados['dt_chegada'])
    ws.cell(row=row_idx, column=3, value=dados['num_nota'])
    ws.cell(row=row_idx, column=4, value=dados['forn'])
    ws.cell(row=row_idx, column=5, value=uf_fornecedor)
    ws.cell(row=row_idx, column=6, value=val_nota).number_format = 'R$ #,##0.00'
    ws.cell(row=row_idx, column=7, value=tipo_frete)
    ws.cell(row=row_idx, column=8, value=frete_combinado_perc).number_format = '0.0%'
    
    ws.cell(row=row_idx, column=10, value=val_blog).number_format = 'R$ #,##0.00'
    ws.cell(row=row_idx, column=11, value=val_terceiro if val_terceiro > 0 else 0).number_format = 'R$ #,##0.00'
    ws.cell(row=row_idx, column=12, value=receita).number_format = 'R$ #,##0.00'
    
    fill_azul = PatternFill(start_color="B8CCE4", end_color="B8CCE4", fill_type="solid")
    fill_laranja = PatternFill(start_color="FCD5B4", end_color="FCD5B4", fill_type="solid")
    
    ws.cell(row=row_idx, column=11).fill = fill_laranja
    ws.cell(row=row_idx, column=12).fill = fill_azul
    
    # === AQUI ESTÁ O BLOQUEIO DA TELA PRETA (PermissionError) ===
    try:
        wb.save(arquivo_blog)
    except PermissionError:
        messagebox.showerror(
            "Aviso de Segurança", 
            f"A planilha de fretes ({os.path.basename(arquivo_blog)}) está aberta no Excel!\n\nPor favor, feche o Excel e audite a nota novamente para guardar os dados da B-LOG."
        )

def salvar_edicao_precos(dados):
    """
    Atualiza SOMENTE VENDA (R$), PRAZO (R$) e MKP REAL no xlsx existente.
    - Salva no MESMO arquivo (sem criar novo)
    - NÃO toca na planilha BLOG
    - Regenera o espelho HTML
    """
    caminho = dados.get('arquivo_aberto_atual')
    if not caminho or not os.path.exists(caminho):
        return False, "Nenhum arquivo aberto. Abra uma precificação salva antes de editar."

    # Índice de preços por código do produto
    precos = {}
    for item in dados['itens']:
        cod = str(item.get('Código', '')).strip()
        if cod:
            precos[cod] = {
                'VENDA (R$)': item.get('VENDA (R$)', ''),
                'PRAZO (R$)': item.get('PRAZO (R$)', ''),
                'MKP REAL':   item.get('MKP REAL', ''),
            }

    # Abrir xlsx e localizar colunas
    try:
        wb = openpyxl.load_workbook(caminho)
    except PermissionError:
        return False, f"O arquivo está aberto no Excel!\nFeche-o e tente novamente.\n{os.path.basename(caminho)}"

    ws = wb['Precificacao']
    header = {ws.cell(row=1, column=c).value: c for c in range(1, ws.max_column + 1)}

    col_cod   = header.get('Código')
    col_venda = header.get('VENDA (R$)')
    col_prazo = header.get('PRAZO (R$)')
    col_mkp   = header.get('MKP REAL')

    if not all([col_cod, col_venda, col_prazo]):
        return False, "Colunas de preço não encontradas no arquivo. Estrutura incompatível."

    # Atualizar somente as células de preço
    for row in range(2, ws.max_row + 1):
        cod = str(ws.cell(row=row, column=col_cod).value or '').strip()
        if cod in precos:
            ws.cell(row=row, column=col_venda).value = precos[cod]['VENDA (R$)']
            ws.cell(row=row, column=col_prazo).value = precos[cod]['PRAZO (R$)']
            if col_mkp:
                ws.cell(row=row, column=col_mkp).value = precos[cod]['MKP REAL']

    try:
        wb.save(caminho)
    except PermissionError:
        return False, f"Não foi possível salvar — o arquivo está aberto no Excel!\n{os.path.basename(caminho)}"

    # Regenerar espelho HTML (sem BLOG)
    _gerar_espelho_html(dados)

    return True, caminho


def _gerar_espelho_html(dados):
    """Gera e abre o espelho HTML executivo de alto padrão."""
    import html as _hm, tempfile as _tmp, webbrowser as _wb
    e = _hm.escape

    forn            = dados.get('forn') or "Sem Fornecedor"
    num_ped_lmp     = re.sub(r'[^\w\-]', '', dados.get('num_pedido', ''))[:20]
    is_pendente     = dados.get('status_pagamento', '') == "PENDENTES"

    forn_safe   = e(forn)
    nota_safe   = e(dados.get('num_nota', ''))
    frete_safe  = e(dados.get('tipo_frete', ''))
    pag_safe    = e(dados.get('status_pagamento', ''))
    pedido_safe = e(num_ped_lmp)
    regime_safe = e(dados.get('regime', ''))

    # Totais financeiros — vêm do dados ou são extraídos do resumo_texto
    resumo_txt = dados.get('resumo_texto', '')
    def _parse(pat):
        m = re.search(pat, resumo_txt)
        return converter_moeda_export(m.group(1)) if m else 0.0
    t_prod  = dados.get('total_mercadoria_compra') or _parse(r'Prod:\s*(R\$\s*[\d\.,]+)')
    t_ipi   = dados.get('total_ipi_compra')        or _parse(r'IPI:\s*(R\$\s*[\d\.,]+)')
    t_frete = dados.get('total_frete_compra')       or _parse(r'Frete:\s*(R\$\s*[\d\.,]+)')
    t_inv   = t_prod + t_ipi + t_frete
    if t_inv == 0:
        m = re.search(r'CUSTO TOTAL CARGA:\s*(R\$\s*[\d\.,]+)', resumo_txt)
        if m: t_inv = converter_moeda_export(m.group(1))

    # Formatadores
    def _brl(v):
        s = f"{abs(v):,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.')
        return f"R$ {s}" if v >= 0 else f"-R$ {s}"
    def _pct(v):
        return f"{v:,.2f}".replace(',', 'X').replace('.', ',').replace('X', '.') + "%"

    # KPIs e indicadores
    ignorar = {"NOTA", "EMISSAO", "CHEGADA", "TIPO FRETE", "VLR TERCEIRO"}
    itens = dados.get('itens') or []
    qtd_itens = len(itens)
    mkp_vals = []; venda_vals = []; retorno_total = 0.0

    for item in itens:
        vv = converter_moeda_export(item.get('VENDA (R$)', ''))
        mkp_s = str(item.get('MKP REAL', '0')).replace('%', '').replace(',', '.').strip()
        qtd_v = converter_moeda_export(item.get('Qtd NF', '') or '0')
        qtd_b = converter_moeda_export(item.get('Qtd Bon', '') or '0')
        tot_q = (qtd_v + qtd_b) if (qtd_v + qtd_b) > 0 else qtd_v
        try: mkp_vals.append(float(mkp_s))
        except: pass
        if vv > 0: venda_vals.append(vv)
        retorno_total += vv * (tot_q if tot_q > 0 else 1)

    mkp_medio   = sum(mkp_vals) / len(mkp_vals) if mkp_vals else 0
    mkp_maior   = max(mkp_vals) if mkp_vals else 0
    mkp_menor   = min(mkp_vals) if mkp_vals else 0
    preco_medio = sum(venda_vals) / len(venda_vals) if venda_vals else 0

    # Cabeçalho da tabela
    colunas = [k for k in (itens[0].keys() if itens else []) if k not in ignorar]
    ths = ''.join(
        f'<th{"" if col != "Produto" else " style=\'text-align:left\'"}'
        f'>{e(str(col))}</th>'
        for col in colunas
    ) + '<th>LUCRO UNIT.</th>'

    # Linhas da tabela
    rows = ''
    for idx, item in enumerate(itens):
        bg    = '#FFFFFF' if idx % 2 == 0 else '#F4F6F7'
        vv    = converter_moeda_export(item.get('VENDA (R$)', ''))
        cv    = converter_moeda_export(item.get('NOVO CUSTO', ''))
        lucro = vv - cv
        lc    = '#1E8449' if lucro >= 0 else '#E74C3C'
        rows += f'<tr style="background:{bg}">'
        for col in colunas:
            val = e(str(item.get(col, '')))
            if col == 'Produto':
                rows += f'<td style="text-align:left;font-weight:500;color:#1A2535;min-width:170px">{val}</td>'
            elif col == 'Código':
                rows += f'<td style="text-align:center;color:#5D6D7E;font-size:10px">{val}</td>'
            elif col == 'NOVO CUSTO':
                rows += f'<td style="background:#FFF3E0;color:#D35400;font-weight:700">{val}</td>'
            elif col == 'VENDA (R$)':
                rows += f'<td style="background:#EAFAF1;color:#1E8449;font-weight:700">{val}</td>'
            elif col == 'PRAZO (R$)':
                rows += f'<td style="background:#EBF5FB;color:#1A5276;font-weight:600">{val}</td>'
            elif col == 'MKP REAL':
                try:
                    mn = float(str(item.get(col, '0')).replace('%', '').replace(',', '.').strip())
                    mc = '#1E8449' if mn >= 30 else '#D35400' if mn < 20 else '#1A5276'
                except:
                    mc = '#1A5276'
                rows += f'<td style="font-weight:700;color:{mc}">{val}</td>'
            elif col in ('% IPI', '% Frete'):
                rows += f'<td style="color:#7F8C8D">{val}</td>'
            else:
                rows += f'<td>{val}</td>'
        rows += f'<td style="font-weight:700;color:{lc}">{_brl(lucro)}</td></tr>'

    alerta = ''
    if is_pendente:
        alerta = ('<div style="background:#FADBD8;border-left:6px solid #E74C3C;padding:12px 20px;'
                  'color:#922B21;font-weight:700;font-size:13px;letter-spacing:.4px">'
                  '⚠️ ATENÇÃO: CARGA PENDENTE DE BOLETOS — AUDITORIA INCOMPLETA</div>')
    badge = ('<div style="background:#E74C3C;color:#fff;padding:5px 12px;border-radius:4px;'
             'font-size:11px;font-weight:700">⚠ PENDENTE</div>') if is_pendente else ''

    # CSS separado (sem f-string para não escapar {})
    css = (
        "*{box-sizing:border-box;margin:0;padding:0}"
        "body{font-family:'Segoe UI',Arial,sans-serif;background:#EDF0F5;color:#1A2535;font-size:12px}"
        ".page{max-width:1600px;margin:0 auto}"
        ".hdr{background:linear-gradient(135deg,#1A2535,#2C3E50);padding:16px 24px;display:flex;"
              "align-items:center;justify-content:space-between}"
        ".hdr-t{color:#F1C40F;font-size:17px;font-weight:700;letter-spacing:.5px}"
        ".hdr-s{color:#AEB6BF;font-size:11px;margin-top:3px}"
        ".ibar{background:#1A2535;display:flex;gap:1px}"
        ".ic{flex:1;padding:10px 14px;background:#223048}"
        ".ic-l{color:#7F8C8D;font-size:9px;font-weight:700;text-transform:uppercase;"
              "letter-spacing:.8px;margin-bottom:3px}"
        ".ic-v{color:#EAECEE;font-size:12px;font-weight:600;white-space:nowrap;"
              "overflow:hidden;text-overflow:ellipsis}"
        ".ic-v.hl{color:#F1C40F}"
        ".kbar{display:flex;gap:12px;padding:14px 20px;background:#EDF0F5}"
        ".kc{flex:1;background:#fff;border-radius:8px;padding:14px 16px;"
             "border-top:4px solid #2980B9;box-shadow:0 2px 8px rgba(0,0,0,.07)}"
        ".kc-i{font-size:18px;margin-bottom:4px}"
        ".kc-l{color:#7F8C8D;font-size:9px;font-weight:700;text-transform:uppercase;"
              "letter-spacing:.8px;margin-bottom:4px}"
        ".kc-v{color:#1A2535;font-size:15px;font-weight:700}"
        ".kc-inv{border-top-color:#1A5276;background:linear-gradient(135deg,#1B2E4A,#2471A3)}"
        ".kc-inv .kc-l{color:#AED6F1}"
        ".kc-inv .kc-v{color:#F1C40F;font-size:18px}"
        ".kc-inv .kc-sub{color:#AED6F1;font-size:9px;margin-top:3px}"
        ".twrap{padding:0 20px 12px;background:#fff;margin-bottom:4px}"
        ".stit{padding:14px 0 8px;font-size:11px;font-weight:700;color:#2C3E50;"
               "text-transform:uppercase;letter-spacing:1px;border-bottom:2px solid #2C3E50;margin-bottom:6px}"
        "table{width:100%;border-collapse:collapse;font-size:11px}"
        "thead tr{background:#1A2535}"
        "thead th{color:#fff;padding:9px 7px;text-align:right;font-size:10px;font-weight:700;"
                 "text-transform:uppercase;letter-spacing:.3px;white-space:nowrap;"
                 "border-right:1px solid #2C3E50}"
        "thead th:first-child{text-align:center}"
        "tbody td{padding:7px 7px;border-bottom:1px solid #ECF0F1;text-align:right;"
                  "vertical-align:middle;border-right:1px solid #ECF0F1}"
        "tbody tr:hover{background:#EBF5FB!important}"
        ".foot{background:linear-gradient(135deg,#1A2535,#2C3E50);margin:0 20px;"
               "border-radius:0 0 8px 8px;padding:14px 20px;display:flex;align-items:center;gap:14px}"
        ".fi{text-align:center;flex:1}"
        ".fi-l{color:#95A5A6;font-size:9px;text-transform:uppercase;letter-spacing:.8px;font-weight:600}"
        ".fi-v{color:#EAECEE;font-size:13px;font-weight:700;margin-top:2px}"
        ".fi-sep{color:#5D6D7E;font-size:22px;font-weight:300;padding:0 4px}"
        ".fi-tot{flex:2;background:rgba(255,255,255,.08);border-radius:6px;padding:10px 16px;"
                "border:1px solid rgba(241,196,15,.4)}"
        ".fi-tot .fi-l{color:#F1C40F}"
        ".fi-tot .fi-v{color:#F1C40F;font-size:18px}"
        ".ibar2{display:flex;gap:12px;padding:12px 20px 20px;background:#EDF0F5}"
        ".ind{flex:1;background:#fff;border-radius:8px;padding:12px 14px;"
              "border-left:4px solid #2471A3;box-shadow:0 2px 6px rgba(0,0,0,.06)}"
        ".ind-l{color:#7F8C8D;font-size:9px;font-weight:700;text-transform:uppercase;"
               "letter-spacing:.8px;margin-bottom:4px}"
        ".ind-v{font-size:15px;font-weight:700;color:#1A2535}"
        ".ind-dst{border-left-color:#F1C40F}"
        ".ind-dst .ind-v{color:#D68910}"
        ".nopr{text-align:center;padding:20px;background:#EDF0F5}"
        ".bpr{padding:11px 28px;font-size:14px;cursor:pointer;background:#2471A3;color:#fff;"
              "border:none;border-radius:6px;font-weight:700;box-shadow:0 3px 8px rgba(0,0,0,.15);"
              "letter-spacing:.3px}"
        ".bpr:hover{background:#1A5276}"
        "@page{margin:12mm}"
        "@media print{"
          "body{background:#fff}"
          ".nopr{display:none!important}"
          ".kbar,.ibar2{background:#fff}"
          ".hdr,.ibar,.kc-inv,.foot{-webkit-print-color-adjust:exact;print-color-adjust:exact}"
          "thead tr{-webkit-print-color-adjust:exact;print-color-adjust:exact}"
          "tbody td{-webkit-print-color-adjust:exact;print-color-adjust:exact}"
        "}"
    )

    html = f"""<!DOCTYPE html>
<html lang="pt-BR">
<head>
<meta charset="utf-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Espelho — {forn_safe}</title>
<style>{css}</style>
</head>
<body>
<div class="page">

<div class="hdr">
  <div>
    <div class="hdr-t">ESPELHO DE PRECIFICAÇÃO E VENDAS</div>
    <div class="hdr-s">BRUNO ELETROMÓVEIS &nbsp;·&nbsp; EDIÇÃO DE PREÇOS</div>
  </div>
  {badge}
</div>

{alerta}

<div class="ibar">
  <div class="ic"><div class="ic-l">Fornecedor</div><div class="ic-v hl">{forn_safe}</div></div>
  <div class="ic"><div class="ic-l">Pedido FDC</div><div class="ic-v">{pedido_safe or '—'}</div></div>
  <div class="ic"><div class="ic-l">Nota Fiscal</div><div class="ic-v">{nota_safe or '—'}</div></div>
  <div class="ic"><div class="ic-l">Regime</div><div class="ic-v">{regime_safe}</div></div>
  <div class="ic"><div class="ic-l">Tipo de Frete</div><div class="ic-v">{frete_safe}</div></div>
  <div class="ic"><div class="ic-l">Pagamento</div><div class="ic-v">{pag_safe}</div></div>
  <div class="ic"><div class="ic-l">Atualizado em</div><div class="ic-v">{datetime.now().strftime('%d/%m/%Y, %H:%M')}</div></div>
</div>

<div class="kbar">
  <div class="kc">
    <div class="kc-i">\U0001f4e6</div>
    <div class="kc-l">Qtd. de Itens</div>
    <div class="kc-v">{qtd_itens} PRODUTOS</div>
  </div>
  <div class="kc kc-inv">
    <div class="kc-i" style="color:#AED6F1">\U0001f4b0</div>
    <div class="kc-l">Investimento Total</div>
    <div class="kc-v">{_brl(t_inv)}</div>
    <div class="kc-sub">CUSTO TOTAL DA CARGA</div>
  </div>
  <div class="kc">
    <div class="kc-i">\U0001f4cb</div>
    <div class="kc-l">IPI Total</div>
    <div class="kc-v">{_brl(t_ipi)}</div>
  </div>
  <div class="kc">
    <div class="kc-i">\U0001f69a</div>
    <div class="kc-l">Frete Total</div>
    <div class="kc-v">{_brl(t_frete)}</div>
  </div>
  <div class="kc">
    <div class="kc-i">\U0001f4c8</div>
    <div class="kc-l">Retorno Potencial</div>
    <div class="kc-v">{_brl(retorno_total)}</div>
  </div>
</div>

<div class="twrap">
  <div class="stit">▸ DETALHAMENTO DOS PRODUTOS</div>
  <table>
    <thead><tr>{ths}</tr></thead>
    <tbody>{rows}</tbody>
  </table>
</div>

<div class="foot">
  <div class="fi"><div class="fi-l">\U0001f4e6 Produtos</div><div class="fi-v">{_brl(t_prod)}</div></div>
  <div class="fi-sep">+</div>
  <div class="fi"><div class="fi-l">\U0001f4cb IPI</div><div class="fi-v">{_brl(t_ipi)}</div></div>
  <div class="fi-sep">+</div>
  <div class="fi"><div class="fi-l">\U0001f69a Frete</div><div class="fi-v">{_brl(t_frete)}</div></div>
  <div class="fi-sep">=</div>
  <div class="fi fi-tot">
    <div class="fi-l">INVESTIMENTO TOTAL DA CARGA</div>
    <div class="fi-v">{_brl(t_inv)}</div>
  </div>
</div>

<div class="ibar2">
  <div class="ind">
    <div class="ind-l">Markup Médio</div>
    <div class="ind-v">{_pct(mkp_medio)}</div>
  </div>
  <div class="ind">
    <div class="ind-l">Preço Médio de Venda</div>
    <div class="ind-v">{_brl(preco_medio)}</div>
  </div>
  <div class="ind">
    <div class="ind-l">Maior Markup</div>
    <div class="ind-v" style="color:#1E8449">{_pct(mkp_maior)}</div>
  </div>
  <div class="ind">
    <div class="ind-l">Menor Markup</div>
    <div class="ind-v" style="color:#D35400">{_pct(mkp_menor)}</div>
  </div>
  <div class="ind ind-dst">
    <div class="ind-l">Investimento Total</div>
    <div class="ind-v">{_brl(t_inv)}</div>
  </div>
</div>

<div class="nopr">
  <button onclick="window.print()" class="bpr">\U0001f5a8️ ENVIAR PARA IMPRESSORA</button>
</div>

</div>
</body>
</html>"""

    p_html = os.path.join(_tmp.gettempdir(), "espelho_impressao_temp.html")
    with open(p_html, 'w', encoding='utf-8') as f:
        f.write(html)
    _wb.open('file://' + os.path.realpath(p_html))